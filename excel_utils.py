from openpyxl import load_workbook, Workbook
import os

EXCEL_FILE = "actors.xlsx"

def ensure_excel_file():
    if not os.path.exists(EXCEL_FILE):
        wb = Workbook()
        ws = wb.active
        ws.append(["Name", "Phone Number", "Age", "Gender", "Paid"])
        wb.save(EXCEL_FILE)

def add_actor_to_excel(name, phone, age, gender, paid=False):
    wb = load_workbook(EXCEL_FILE)
    ws = wb.active
    ws.append([name, phone, age, gender, 1 if paid else 0])
    wb.save(EXCEL_FILE)

def search_actors_in_excel(gender=None, age_min=None, age_max=None):
    wb = load_workbook(EXCEL_FILE)
    ws = wb.active
    results = []

    for row in ws.iter_rows(min_row=2, values_only=True):
        name, phone, age, gdr, paid = row
        match = True

        if gender and gdr != gender:
            match = False
        if age_min and age < age_min:
            match = False
        if age_max and age > age_max:
            match = False

        if match:
            results.append({
                "name": name,
                "phone": phone,
                "age": age,
                "gender": gdr,
                "paid": bool(paid)
            })

    return results
